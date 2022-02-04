from openpyxl import load_workbook
from PyQt5.QtWidgets import QComboBox, QListView
from PyQt5.QtCore import QStringListModel


def load_excel() -> list:
    """Загрузка данных из файла Excel, fname - полный путь к файлу"""
    # ссылка на лист в таблице
    fname = 'data.xlsx'
    wb = load_workbook(fname)
    sheet = wb.get_sheet_by_name(name = 'climat')
    # количество заполненных строк
    max_row = sheet.max_row
    # чтение данных
    data_cities = []
    for i in range(6, max_row+1):
        row_citi = []
        for j in range(1, 21):
            row_citi.append(sheet.cell(row=i, column=j).value)
        data_cities.append(row_citi)
    return data_cities


def load_windows_koef() -> dict:
    """Загрузка характеристик окон из файла Excel, fname - полный путь к файлу
    :return - словарь key = тип стеклопакета, list = коэффициенты"""
    fname = 'data.xlsx'
    # ссылка на лист в таблице
    wb = load_workbook(fname)
    sheet = wb.get_sheet_by_name(name = 'windows')
    # количество заполненных строк
    max_row = sheet.max_row
    # чтение данных
    data_windows = dict()
    for i in range(1, max_row+1):
        koef = []
        for j in range(2, 4):
            koef.append(sheet.cell(row=i, column=j).value)
        key = f'{sheet.cell(row=i, column=1).value} (τ={koef[0]} g={koef[1]})'
        data_windows[key] = koef
    return data_windows


def load_solar_radiation() -> dict:
    """Загрузка солнечной радиации из файла Excel, fname - полный путь к файлу
    :return - словарь key = название города -> key - название параметра -> list - значения"""
    fname = 'data.xlsx'
    data = dict()
    wb = load_workbook(fname)
    # загрузка солнечной радиации для городов
    sheet = wb.get_sheet_by_name(name = 'solar')
    max_row = sheet.max_row
    for i in range(1, max_row + 1):
        citi = sheet.cell(row=i, column=1).value
        if citi not in data:
            parameters = dict()
        key = sheet.cell(row=i, column=2).value
        parametr = []
        for j in range(3, 15):
            p = sheet.cell(row=i, column=j).value
            if p == None:
                p = 0
            parametr.append(p)
        parameters[key] = parametr
        data[citi] = parameters
    return data


def load_orientation_coef() -> dict:
    """Загрузка коэфициентов ориентации из файла Excel, fname - полный путь к файлу
    :return - словарь key - ориентация -> key - широта -> list - значения"""
    fname = 'data.xlsx'
    data = dict()
    wb = load_workbook(fname)
    # загрузка солнечной радиации для городов
    sheet = wb.get_sheet_by_name(name = 'coef')
    max_row = sheet.max_row
    for i in range(1, max_row + 1):
        azimut = sheet.cell(row=i, column=1).value
        if azimut not in data:
            parameters = dict()
        key = sheet.cell(row=i, column=2).value
        parametr = []
        for j in range(3, 15):
            p = sheet.cell(row=i, column=j).value
            if p == None:
                p = 0
            parametr.append(p)
        parameters[key] = parametr
        data[azimut] = parameters
    return data


def to_dot(s: str) -> str:
    if ',' in s:
        s = s.replace(',', '.')
    return s


def to_float(s: str) -> float:
    """Конвертация строки в вещественный тип данных
    :param s - строка
    :return тип флоат"""
    try:
        x = float(to_dot(s))
    except:
        print('Некореткно указано значение')
        x = 0.0
    return x

def to_int(s: str) -> int:
    """Конвертация строки в вещественный тип данных
    :param s - строка
    :return переменная типа int"""
    try:
        x = int(s)
    except:
        print('Некореткно указано значение')
        x = 0
    return x

def get_string_index(index: int) -> str:
    """Расчет нижнего индекса для значения
    :return - строковое представление индекса"""
    list_underline = ['₀', '₁', '₂', '₃', '₄', '₅', '₆', '₇', '₈', '₉']
    s_index = ''
    if type(index) == int:
        letters = str(index)
        for s in letters:
            try:
                i = int(s)
            except:
                print('Ошибка при конвертации номера слоя')
                i = -1
            if i > -1:
                s_index += list_underline[i]
    return s_index


def r_unit() -> str:
    """Возвращает единицу измерения сопротивления теплопередаче"""
    return 'м²·ºС/Вт'


def l_unit() -> str:
    """Возвращает единицу измерения коэффициент теплопроводности"""
    return 'Вт/(м·ºС)'


def alfa_unit() -> str:
    """Возвращает единицу измерения коэффициента теплоотдачи"""
    return 'Вт/(м²·ºС/Вт)'


class MyCombo(QComboBox):
    def __init__(self, elements: list):
        super().__init__()
        self.setModel(QStringListModel(elements))
        listview = QListView()
        listview.setWordWrap(True)
        self.setView(listview)
